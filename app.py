"""Daily Report Rogério Salvador — Dashboard de carteira sem login"""
import json, os
from pathlib import Path
from flask import Flask, jsonify, request, render_template
from io import BytesIO

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

app = Flask(__name__)
BASE_DIR = Path(__file__).parent
DATA_FILE = BASE_DIR / "data" / "relatorio.json"
GRUPOS = ["NAGUMO", "FESTVAL", "JACOMAR"]
PERIODOS = ["D-1", "D-7", "D-15", "D-21", "MTD", "CONSOLIDADO"]


def load_data():
    if DATA_FILE.exists():
        return json.loads(DATA_FILE.read_text(encoding="utf-8"))
    return {}


def semaforo(kpi, val):
    if val is None:
        return "gray"
    try:
        v = float(val)
    except Exception:
        return "gray"
    bad = {"cancel_pct": (5, 8), "ruptura_pct": (10, 25), "er_pct": (10, 20)}
    good = {"ating_pct": (90, 80), "nps": (70, 50), "sla_pct": (90, 70), "online_pct": (95, 80)}
    if kpi in bad:
        lo, hi = bad[kpi]
        return "green" if v < lo else ("yellow" if v < hi else "red")
    if kpi in good:
        hi, lo = good[kpi]
        return "green" if v >= hi else ("yellow" if v >= lo else "red")
    return "gray"


def fmt_brl(v):
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"


def fmt_pct(v, d=1):
    try:
        return f"{float(v):.{d}f}%"
    except Exception:
        return "—"


@app.route("/")
def index():
    data = load_data()
    lojas_json = json.dumps(data.get("lojas", {})) if data else "{}"
    return render_template(
        "index.html",
        data=data,
        grupos=GRUPOS,
        periodos=PERIODOS,
        lojas_json=lojas_json,
        fmt_brl=fmt_brl,
        fmt_pct=fmt_pct,
        semaforo=semaforo,
    )


@app.route("/atualizar", methods=["POST"])
def atualizar():
    try:
        payload = request.get_json(silent=True)
        if not payload:
            return jsonify({"error": "JSON inválido"}), 400
        DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
        DATA_FILE.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        return jsonify({"status": "ok", "message": "Dados atualizados"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/dados")
def api_dados():
    data = load_data()
    if not data:
        return jsonify({"error": "Sem dados"}), 404
    return jsonify(data)


@app.route("/download/excel/<grupo>")
def download_excel(grupo):
    if grupo not in GRUPOS:
        return f"Grupo inválido: {grupo}", 400
    if not EXCEL_OK:
        return "openpyxl não instalado", 500
    data = load_data() or {}
    lojas = data.get("lojas", {}).get(grupo, {})
    wb = openpyxl.Workbook()
    cols = ["Loja", "GMV", "Pedidos", "Ating%", "AOV", "Cancel%", "Ruptura%", "ER%", "NPS", "SLA%", "Online%", "GMV Rupt", "GMV Recup", "NSU%"]
    keys = ["nome", "gmv", "pedidos", "ating_pct", "aov", "cancel_pct", "ruptura_pct", "er_pct", "nps", "sla_pct", "online_pct", "gmv_rupt", "gmv_recup", "nsu_pct"]
    for periodo in PERIODOS:
        ws = wb.create_sheet(title=periodo)
        for col, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="EA1D2C")
            cell.alignment = Alignment(horizontal="center")
        for ri, loja in enumerate(lojas.get(periodo, []), 2):
            for ci, k in enumerate(keys, 1):
                ws.cell(row=ri, column=ci, value=loja.get(k, 0))
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import Response
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={grupo.lower()}_carteira.xlsx"},
    )



@app.route("/health")
def health():
    return jsonify({"status": "ok"}), 200


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)
